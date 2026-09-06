import os
from datetime import datetime
from typing import List, Optional, Sequence, Union
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.models.schemas import LeadRecord, SheetLeadRecord, normalize_call_status
from app.config import settings

CALL_STATUS_OPTIONS = [
    "Pending",
    "Interested",
    "Not Interested",
    "Not Reachable",
    "Other",
]

def generate_leads_excel(
    leads: Sequence[Union[LeadRecord, SheetLeadRecord]],
    location: str,
    job_id: str,
    output_dir: Optional[str] = None,
) -> str:
    wb = Workbook()
    ws = wb.active

    safe_location = "".join(c for c in location if c.isalnum() or c in (" ", "_", "-")).strip()
    sheet_title = f"Leads - {safe_location}"[:31]
    ws.title = sheet_title
    ws.views.sheetView[0].showGridLines = True

    # Styling Palettes
    navy_dark = "1E293B"       # Primary header
    navy_title = "0F172A"      # Main title banner
    text_white = "FFFFFF"
    sub_text = "94A3B8"
    zebra_even = "F8FAFC"
    border_gray = "CBD5E1"
    link_blue = "2563EB"

    thin_side = Side(style="thin", color=border_gray)
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 1. Top Title Banner (Rows 1 & 2)
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"LOCALLEADPULSE B2B SALES PIPELINE — {location.upper()}"
    title_cell.font = Font(name="Segoe UI", size=14, bold=True, color=text_white)
    title_cell.fill = PatternFill(start_color=navy_title, end_color=navy_title, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    sub_cell = ws["A2"]
    now_str = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    sub_cell.value = f"Generated: {now_str} | Total Qualified Leads: {len(leads)} | Verified No Standalone Website"
    sub_cell.font = Font(name="Segoe UI", size=9, italic=True, color=sub_text)
    sub_cell.fill = PatternFill(start_color=navy_title, end_color=navy_title, fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 10  # Spacer row

    # 2. Table Headers (Row 4) - Pitch Angle column completely removed
    headers = [
        "ID",
        "Shop / Business Name",
        "Category",
        "Contact Number",
        "Full Physical Address",
        "Area / Landmark",
        "Google Maps URL",
        "Web Search Verification Status",
        "Call Status",
        "Lead Identified Date",
    ]

    header_font = Font(name="Segoe UI", size=10, bold=True, color=text_white)
    header_fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_row = 4
    ws.row_dimensions[header_row].height = 28

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border

    # 3. Data Rows
    body_font = Font(name="Segoe UI", size=9.5)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    link_font = Font(name="Segoe UI", size=9.5, color=link_blue, underline="single")

    for row_offset, lead in enumerate(leads):
        row_num = header_row + 1 + row_offset
        ws.row_dimensions[row_num].height = 24
        is_even = (row_offset % 2 == 1)
        fill = PatternFill(start_color=zebra_even, end_color=zebra_even, fill_type="solid") if is_even else None

        row_data = [
            (lead.id, center_align, None),
            (lead.name, left_align, Font(name="Segoe UI", size=9.5, bold=True)),
            (lead.category, center_align, None),
            (lead.phone, center_align, None),
            (lead.address, left_align, None),
            (lead.area or location, center_align, None),
            ("Open Google Map", center_align, link_font),  # Clean text label
            (lead.verification_status or "No Standalone Website Found", center_align, None),
            (normalize_call_status(lead.call_status), center_align, None),
            (lead.date_identified, center_align, None),
        ]

        for col_idx, (val, align, custom_font) in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = custom_font or body_font
            cell.alignment = align
            cell.border = cell_border
            if fill:
                cell.fill = fill

            # Hyperlink Google Maps
            if col_idx == 7 and lead.maps_url:
                cell.hyperlink = lead.maps_url

    # 4. Native Excel Dropdown Validation on Call Status (Column I)
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(CALL_STATUS_OPTIONS)}"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    max_data_row = max(header_row + len(leads), header_row + 50)
    dv.add(f"I{header_row + 1}:I{max_data_row}")

    # 5. Enable Excel Sort & Filter Dropdown Arrows across all columns
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row + len(leads)}"

    # 6. Auto-fit column widths
    col_widths = {
        "A": 10,  # ID
        "B": 28,  # Name
        "C": 14,  # Category
        "D": 18,  # Phone
        "E": 40,  # Address
        "F": 18,  # Area
        "G": 20,  # Map link
        "H": 28,  # Status
        "I": 18,  # Call Status
        "J": 20,  # Date
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A5"

    export_dir = output_dir or settings.EXPORTS_DIR
    os.makedirs(export_dir, exist_ok=True)
    file_path = os.path.join(export_dir, f"leads_{job_id}.xlsx")
    wb.save(file_path)
    return file_path