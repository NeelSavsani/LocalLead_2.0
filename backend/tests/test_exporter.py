import os
import openpyxl
from app.models.schemas import LeadRecord
from app.services.excel_exporter import generate_leads_excel, CALL_STATUS_OPTIONS


def test_generate_leads_excel_creates_valid_styled_file(tmp_path):
    leads = [
        LeadRecord(
            id="LLP-001",
            name="Maruti Motors & Car Care",
            category="Auto Garage",
            phone="+91 98250 14820",
            address="Plot 12, GIDC Electronic Estate, Gandhinagar",
            area="Sector 25",
            maps_url="https://maps.google.com/?q=Maruti",
            pitch_angle="WhatsApp service booking & maintenance reminders",
        ),
        LeadRecord(
            id="LLP-002",
            name="Shree Ram Auto Garage",
            category="Auto Garage",
            phone="+91 94265 88310",
            address="Pramukh Arcade, Kudasan, Gandhinagar",
            area="Kudasan",
            maps_url="https://maps.google.com/?q=ShreeRam",
            pitch_angle="Online service slot reservation portal",
        ),
    ]

    job_id = "test-job-999"
    output_dir = str(tmp_path)

    file_path = generate_leads_excel(
        job_id=job_id,
        location="Gandhinagar, Gujarat",
        leads=leads,
        output_dir=output_dir,
    )

    # 1. File verification
    assert os.path.exists(file_path)
    assert os.path.getsize(file_path) > 0

    # 2. Inspect with openpyxl
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    assert "Leads - " in sheet.title

    # 3. Banner verification
    assert "LOCALLEADPULSE B2B SALES PIPELINE" in sheet["A1"].value
    assert "Total Qualified Leads: 2" in sheet["A2"].value

    # 4. Header verification (Row 4)
    expected_headers = [
        "ID",
        "Shop / Business Name",
        "Category",
        "Contact Number",
        "Full Physical Address",
        "Area / Landmark",
        "Google Maps URL",
        "Web Search Verification Status",
        "Call Status",
        "Pitch Angle / Opportunity",
        "Lead Identified Date",
    ]
    for col_idx, expected in enumerate(expected_headers, start=1):
        cell = sheet.cell(row=4, column=col_idx)
        assert cell.value == expected
        # Dark navy fill
        assert cell.fill.start_color.rgb in ["001E293B", "1E293B"]
        # Bold white text
        assert cell.font.bold is True
        assert cell.font.color.rgb in ["00FFFFFF", "FFFFFF"]

    # 5. Data rows verification
    row5_id = sheet.cell(row=5, column=1).value
    row5_name = sheet.cell(row=5, column=2).value
    row5_phone = sheet.cell(row=5, column=4).value

    assert row5_id == "LLP-001"
    assert row5_name == "Maruti Motors & Car Care"
    assert row5_phone == "+91 98250 14820"

    row6_id = sheet.cell(row=6, column=1).value
    assert row6_id == "LLP-002"

    # Zebra striping: Row 5 fill vs Row 6 fill
    row5_fill = sheet.cell(row=5, column=1).fill.start_color.rgb
    row6_fill = sheet.cell(row=6, column=1).fill.start_color.rgb
    assert row5_fill in ["00FFFFFF", "FFFFFF"]
    assert row6_fill in ["00F8FAFC", "F8FAFC"]

    # 6. Data Validation dropdown for Call Status
    assert len(sheet.data_validations.dataValidation) > 0
    dv = sheet.data_validations.dataValidation[0]
    for opt in CALL_STATUS_OPTIONS:
        assert opt in dv.formula1

    # 7. Freeze Panes
    assert sheet.freeze_panes == "A5"
